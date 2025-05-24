import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
import io
import streamlit as st
from typing import Dict, List, Optional, Tuple, Any
import xlsxwriter


class ExcelHandler:
    """Utility class for handling Excel file operations"""

    @staticmethod
    def read_excel_file(uploaded_file, skip_rows: int = 0) -> Dict[str, pd.DataFrame]:
        """
        Read an Excel file and return a dictionary of DataFrames (one per sheet)

        Args:
            uploaded_file: Streamlit uploaded file object
            skip_rows: Number of rows to skip before header row

        Returns:
            Dictionary with sheet names as keys and DataFrames as values
        """
        try:
            # Read all sheets into a dictionary
            sheets_dict = pd.read_excel(
                uploaded_file, sheet_name=None, engine="openpyxl", skiprows=skip_rows
            )

            # Clean up the data - remove completely empty rows/columns
            cleaned_sheets = {}
            for sheet_name, df in sheets_dict.items():
                # Remove rows that are completely empty
                df_cleaned = df.dropna(how="all")
                # Remove columns that are completely empty
                df_cleaned = df_cleaned.dropna(axis=1, how="all")
                cleaned_sheets[sheet_name] = df_cleaned

            return cleaned_sheets

        except Exception as e:
            st.error(f"Error reading Excel file: {str(e)}")
            return {}

    @staticmethod
    def get_sheet_names(uploaded_file) -> List[str]:
        """
        Get list of sheet names from an Excel file

        Args:
            uploaded_file: Streamlit uploaded file object

        Returns:
            List of sheet names
        """
        try:
            workbook = openpyxl.load_workbook(uploaded_file, read_only=True)
            sheet_names = workbook.sheetnames
            workbook.close()
            return sheet_names
        except Exception as e:
            st.error(f"Error reading sheet names: {str(e)}")
            return []

    @staticmethod
    def get_column_names(df: pd.DataFrame) -> List[str]:
        """
        Get list of column names from a DataFrame

        Args:
            df: pandas DataFrame

        Returns:
            List of column names
        """
        return df.columns.tolist()

    @staticmethod
    def validate_file(uploaded_file) -> Tuple[bool, str]:
        """
        Validate if uploaded file is a valid Excel file

        Args:
            uploaded_file: Streamlit uploaded file object

        Returns:
            Tuple of (is_valid, error_message)
        """
        if uploaded_file is None:
            return False, "No file uploaded"

        # Check file extension
        if not uploaded_file.name.lower().endswith((".xlsx", ".xls")):
            return False, "File must be an Excel file (.xlsx or .xls)"

        # Check file size (limit to 50MB)
        if uploaded_file.size > 50 * 1024 * 1024:
            return False, "File size must be less than 50MB"

        try:
            # Try to read the file to validate it's not corrupted
            workbook = openpyxl.load_workbook(uploaded_file, read_only=True)
            workbook.close()
            return True, "File is valid"
        except Exception as e:
            return False, f"Invalid Excel file: {str(e)}"

    @staticmethod
    def create_download_excel(
        data: Dict[str, pd.DataFrame], filename: str = "output.xlsx"
    ) -> bytes:
        """
        Create an Excel file with multiple sheets for download

        Args:
            data: Dictionary with sheet names as keys and DataFrames as values
            filename: Name for the output file

        Returns:
            Bytes object of the Excel file
        """
        output = io.BytesIO()

        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
            for sheet_name, df in data.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

                # Get the workbook and worksheet objects
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]

                # Auto-adjust column widths
                for i, col in enumerate(df.columns):
                    # Calculate the maximum width needed
                    max_len = max(
                        df[col].astype(str).str.len().max(),  # Max length in column
                        len(str(col)),  # Length of column name
                    )
                    # Set column width (with some padding)
                    worksheet.set_column(i, i, min(max_len + 2, 50))

        output.seek(0)
        return output.getvalue()

    @staticmethod
    def create_highlighted_excel(
        df: pd.DataFrame,
        highlight_rows: List[int],
        sheet_name: str = "Results",
        highlight_color: str = "FFFF00",
    ) -> bytes:
        """
        Create an Excel file with highlighted rows

        Args:
            df: DataFrame to write
            highlight_rows: List of row indices to highlight
            sheet_name: Name of the sheet
            highlight_color: Hex color code for highlighting

        Returns:
            Bytes object of the Excel file
        """
        output = io.BytesIO()

        # Create workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name

        # Write headers
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_title
            cell.font = Font(bold=True)

        # Write data
        for row_num, row_data in enumerate(df.values, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

                # Highlight if row is in highlight list
                if (
                    row_num - 2
                ) in highlight_rows:  # -2 because of header and 0-indexing
                    cell.fill = PatternFill(
                        start_color=highlight_color,
                        end_color=highlight_color,
                        fill_type="solid",
                    )

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Save to bytes
        workbook.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def preview_dataframe(df: pd.DataFrame, max_rows: int = 10) -> pd.DataFrame:
        """
        Create a preview of the DataFrame for display

        Args:
            df: DataFrame to preview
            max_rows: Maximum number of rows to show

        Returns:
            Truncated DataFrame for preview
        """
        if len(df) > max_rows:
            preview_df = df.head(max_rows).copy()
            # Add a note about truncation
            return preview_df
        return df

    @staticmethod
    def preview_raw_excel(uploaded_file, max_rows: int = 10) -> pd.DataFrame:
        """
        Get a raw preview of the Excel file without any processing

        Args:
            uploaded_file: Streamlit uploaded file object
            max_rows: Maximum number of rows to preview

        Returns:
            DataFrame with raw data preview
        """
        try:
            # Read without any processing - just raw data
            preview_df = pd.read_excel(
                uploaded_file,
                sheet_name=0,  # First sheet only
                engine="openpyxl",
                header=None,  # Don't treat any row as header
                nrows=max_rows,
            )
            return preview_df
        except Exception as e:
            st.error(f"Error previewing file: {str(e)}")
            return pd.DataFrame()
        """
        Get summary information about a DataFrame
        
        Args:
            df: DataFrame to analyze
            
        Returns:
            Dictionary with DataFrame statistics
        """
        return {
            "rows": len(df),
            "columns": len(df.columns),
            "column_names": df.columns.tolist(),
            "data_types": df.dtypes.to_dict(),
            "memory_usage": df.memory_usage(deep=True).sum(),
            "null_counts": df.isnull().sum().to_dict(),
        }

    @staticmethod
    def get_dataframe_info(df):
        if df is None:
            return "No DataFrame provided."

        info = {
            "columns": list(df.columns),
            "dtypes": df.dtypes.to_dict(),
            "shape": df.shape,
            "rows": len(df),
            "null_counts": df.isnull().sum().to_dict(),
            "memory_usage": df.memory_usage(deep=True).sum(),  # in bytes
        }

        return info
